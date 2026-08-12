#!/usr/bin/env python3
"""Build the Google-Sheets-ready workbook from out/messages.csv + out/stats.json + out/playbook.json.

Sheets:
  Summary   – headline numbers, disposition mix, automation ceiling, per-group
  Playbook  – intent -> disposition -> action -> reply template -> owner -> phase
  Messages  – one row per message + the bot's decision (the living data sheet)
  How-to    – how to keep adding messages + import to Google Sheets
"""
import csv, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "out")
messages = list(csv.reader(open(os.path.join(OUT, "messages.csv"), newline="", encoding="utf-8")))
stats = json.load(open(os.path.join(OUT, "stats.json")))
pb = json.load(open(os.path.join(OUT, "playbook.json")))

def _load_csv(name):
    p = os.path.join(OUT, name)
    return list(csv.reader(open(p, newline="", encoding="utf-8"))) if os.path.exists(p) else None
def _load_json(name):
    p = os.path.join(OUT, name)
    return json.load(open(p)) if os.path.exists(p) else None

handling = _load_csv("handling.csv")            # learned Q->A pairs (may be None)
reply_library = _load_json("reply-library.json")  # learned real reply phrasings

HEAD = messages[0]
ROWS = messages[1:]

# ── palette ──────────────────────────────────────────────────────────────
INK = "1F2937"; MUTED = "6B7280"; LINE = "E5E7EB"
NAVY = "0F2A43"; ACCENT = "2563EB"
FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_BAND = PatternFill("solid", fgColor="F3F4F6")
FILL_TITLE = PatternFill("solid", fgColor="0B3D5C")
ACTION_FILL = {
    "AUTO_REPLY":      "DCFCE7",  # green  – bot answers
    "AUTO_ASK_ID":     "DCFCE7",
    "OUTBOUND_POST":   "D1FAE5",
    "CONSOLE_TASK":    "DBEAFE",  # blue   – to console
    "LAB_ESCALATION":  "FEF3C7",  # amber  – to lab
    "HUMAN_ESCALATION":"FEE2E2",  # red    – to human
    "LLM_REVIEW":      "EDE9FE",  # purple – LLM layer
    "IGNORE":          "F3F4F6",  # grey
}
CONF_FILL = {"HIGH": "DCFCE7", "MED": "FEF3C7", "LOW": "FEE2E2"}
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols, fill=FILL_HEAD, color="FFFFFF"):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=color, size=11)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26

wb = Workbook()

# ══════════════════════════ Summary ═══════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
t = stats["totals"]; c = stats["ceiling"]; disp = stats["dispositions"]
sub = t["substantive"]

ws["A1"] = "LabStack WhatsApp Bot — Traffic & Automation Analysis"
ws["A1"].font = Font(bold=True, size=16, color=NAVY)
ws["A2"] = f"Corpus: 4 partner-ops groups · {t['messages']:,} messages · {sub:,} substantive (non-noise) · {t['withId']:,} carry an order/booking id"
ws["A2"].font = Font(size=10, color=MUTED)

def block(ws, r, title):
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12, color=ACCENT)

# headline tiles
block(ws, 4, "Headline")
tiles = [
    ("Auto-answered by bot today", f"{c['autoAnswerNow_pctOfSubstantive']}%", "status/report/cancel-reason, from LabStack"),
    ("With LLM+context layer", f"{c['autoAnswerWithLLM_pctOfSubstantive']}%", "bare-ids + half of freeform collapse in"),
    ("No human reply needed", f"{c['fullyDeflectable_pctOfSubstantive']}%", "auto-answer + routed to console/outbound"),
    ("Needs a human today", f"{disp.get('HUMAN',0)/sub*100:.1f}%", "escalations + tech issues only"),
]
r = 5
for label, val, sub2 in tiles:
    ws.cell(row=r, column=1, value=label).font = Font(size=10, color=MUTED)
    ws.cell(row=r, column=2, value=val).font = Font(bold=True, size=14, color=NAVY)
    ws.cell(row=r, column=3, value=sub2).font = Font(size=9, color=MUTED)
    r += 1

# disposition mix
r += 1
block(ws, r, "Disposition mix (share of substantive)"); r += 1
ws.cell(row=r, column=1, value="Disposition").font = Font(bold=True)
ws.cell(row=r, column=2, value="Count").font = Font(bold=True)
ws.cell(row=r, column=3, value="Share").font = Font(bold=True)
ws.cell(row=r, column=4, value="What it means").font = Font(bold=True)
style_header(ws, r, 4)
DISP_MEAN = {
    "AUTO_ANSWER": "Bot replies with live status — no human",
    "ROUTE_CONSOLE": "Bot opens an OpsFlow task — team works it in console",
    "NEEDS_LAB": "Serviceability / slot / price — lab data or lab group",
    "OUTBOUND": "LabStack-side status post — bot can auto-generate",
    "HUMAN": "Escalation / tech issue — person needed fast",
    "REVIEW": "Ambiguous / bare-id — the LLM+context layer resolves",
}
r += 1
for k, v in sorted(disp.items(), key=lambda kv: -kv[1]):
    if k == "NOISE":
        continue
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=v)
    ws.cell(row=r, column=3, value=f"{v/sub*100:.1f}%")
    ws.cell(row=r, column=4, value=DISP_MEAN.get(k, ""))
    for cc in range(1, 5):
        ws.cell(row=r, column=cc).border = BORDER
        if r % 2 == 0:
            ws.cell(row=r, column=cc).fill = FILL_BAND
    r += 1

# per-group
r += 1
block(ws, r, "Per group"); r += 1
gh = ["Group", "Active days", "Messages", "Substantive", "With id"]
for i, h in enumerate(gh, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, len(gh))
r += 1
for g in stats["perGroup"]:
    ws.cell(row=r, column=1, value=g["name"])
    ws.cell(row=r, column=2, value=g["days"])
    ws.cell(row=r, column=3, value=g["total"])
    ws.cell(row=r, column=4, value=g["substantive"])
    ws.cell(row=r, column=5, value=g["withId"])
    for cc in range(1, len(gh) + 1):
        ws.cell(row=r, column=cc).border = BORDER
    r += 1

for col, w in {"A": 34, "B": 14, "C": 12, "D": 44, "E": 10}.items():
    ws.column_dimensions[col].width = w

# ══════════════════════════ Playbook ══════════════════════════════════════
ws = wb.create_sheet("Playbook")
ws.sheet_view.showGridLines = False
ws["A1"] = "Intent → Action Playbook (the bot's decision table)"
ws["A1"].font = Font(bold=True, size=14, color=NAVY)
ph = ["Intent", "Disposition", "Action type", "Needs DB lookup", "Phase", "Owner", "What the bot does", "Example reply / note"]
for i, h in enumerate(ph, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(ph))
r = 4
# order by phase then disposition for readability
order = {"P0": 0, "P1": 1, "P2": 2}
for it in sorted(pb["intents"], key=lambda x: (order.get(x["phase"], 9), x["disposition"])):
    if it["intent"] in ("NOISE", "SYSTEM"):
        continue
    vals = [it["intent"], it["disposition"], it["action_type"], "Yes" if it["needs_db_lookup"] else "—",
            it["phase"], it["owner"], it["summary"], it["example_reply"]]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    fill = ACTION_FILL.get(it["action_type"])
    if fill:
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=fill)
    r += 1
for col, w in {"A": 17, "B": 15, "C": 17, "D": 15, "E": 8, "F": 9, "G": 46, "H": 52}.items():
    ws.column_dimensions[col].width = w

# status phrasing tables
r += 1
ws.cell(row=r, column=1, value="Order status → partner-facing phrasing").font = Font(bold=True, color=ACCENT); r += 1
for k, v in pb["orderStatus"].items():
    ws.cell(row=r, column=1, value=k); ws.cell(row=r, column=2, value=v); r += 1
r += 1
ws.cell(row=r, column=1, value="Request status → partner-facing phrasing").font = Font(bold=True, color=ACCENT); r += 1
for k, v in pb["requestStatus"].items():
    ws.cell(row=r, column=1, value=k); ws.cell(row=r, column=2, value=v); r += 1

# ══════════════════════════ Messages ══════════════════════════════════════
ws = wb.create_sheet("Messages")
ws.freeze_panes = "A2"
# nicer header labels
LABELS = {
    "date": "Date", "time": "Time", "group": "Group", "sender": "Sender", "side": "Side",
    "message": "Message", "has_id": "Has id", "ids": "Ids", "intent": "Intent",
    "disposition": "Disposition", "action_type": "Action type", "needs_db_lookup": "Needs DB",
    "confidence": "Confidence", "phase": "Phase", "owner": "Owner",
    "bot_action_or_reply": "Bot action / suggested reply",
}
for i, h in enumerate(HEAD, 1):
    ws.cell(row=1, column=i, value=LABELS.get(h, h))
style_header(ws, 1, len(HEAD))

ci = {h: i for i, h in enumerate(HEAD)}
for ri, row in enumerate(ROWS, start=2):
    for i, v in enumerate(row, 1):
        ws.cell(row=ri, column=i, value=v)
    # colour action_type + confidence
    at = row[ci["action_type"]]
    if at in ACTION_FILL:
        ws.cell(row=ri, column=ci["action_type"] + 1).fill = PatternFill("solid", fgColor=ACTION_FILL[at])
    cf = row[ci["confidence"]]
    if cf in CONF_FILL:
        ws.cell(row=ri, column=ci["confidence"] + 1).fill = PatternFill("solid", fgColor=CONF_FILL[cf])

widths = {"date": 10, "time": 11, "group": 26, "sender": 20, "side": 7, "message": 60,
          "has_id": 7, "ids": 12, "intent": 15, "disposition": 14, "action_type": 17,
          "needs_db_lookup": 9, "confidence": 11, "phase": 7, "owner": 9, "bot_action_or_reply": 50}
for h, w in widths.items():
    ws.column_dimensions[get_column_letter(ci[h] + 1)].width = w
# an Excel table over the data → filter/sort, and easy to append rows to
last = get_column_letter(len(HEAD)) + str(len(ROWS) + 1)
tbl = Table(displayName="Messages", ref=f"A1:{last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
ws.add_table(tbl)

# ══════════════════════════ How-to ════════════════════════════════════════
ws = wb.create_sheet("How-to")
ws.sheet_view.showGridLines = False
ws["A1"] = "How to use & keep growing this sheet"
ws["A1"].font = Font(bold=True, size=14, color=NAVY)
lines = [
    "",
    "WHAT THIS IS",
    "  • Every message from the LabStack partner-ops groups, with the exact action the WhatsApp bot",
    "    would take — classified by the same engine that runs live (dry-run today, sending nothing).",
    "  • 'Messages' is the living data. 'Playbook' is the decision table. 'Summary' is the dashboard.",
    "",
    "COLOUR KEY (Action type)",
    "  • Green  = bot auto-replies (status / report / cancel-reason)   → no human",
    "  • Blue   = routed to an OpsFlow console task                    → team actions it",
    "  • Amber  = serviceability / slot / price                       → lab data or lab group",
    "  • Red    = escalation / tech issue                             → a person, fast",
    "  • Purple = LLM+context layer resolves (bare id / freeform)",
    "  • Grey   = noise / system — ignored",
    "",
    "PUT IT ON GOOGLE SHEETS",
    "  1. Google Sheets → File → Import → Upload → this .xlsx → 'Replace spreadsheet'.",
    "  2. The four tabs come across as-is; 'Messages' keeps its filter.",
    "",
    "KEEP ADDING NEW MESSAGES (from the live dry-run bot)",
    "  • The bot logs every observed message to whatsapp-bot/dry-run.log.jsonl.",
    "  • Run:  node scripts/log-to-csv.mjs   → appends new rows to out/messages.csv (same columns).",
    "  • Then in Google Sheets: File → Import → Upload out/messages.csv → 'Append to current sheet'",
    "    onto the Messages tab. (Or re-run build_xlsx.py to rebuild the whole workbook.)",
    "",
    "TUNE THE BOT FROM THIS SHEET",
    "  • Skim the Intent / Action columns; where a row is wrong, note the correct intent.",
    "  • Those corrections feed the classifier rules (lib/classifier.mjs) and the LLM prompt.",
]
for i, ln in enumerate(lines, start=3):
    cell = ws.cell(row=i, column=1, value=ln)
    if ln.isupper() and ln.strip():
        cell.font = Font(bold=True, color=ACCENT, size=11)
    else:
        cell.font = Font(size=10, color=INK)
ws.column_dimensions["A"].width = 110

# ══════════════════════════ Handling (learned Q→A) ════════════════════════
RES_FILL = {
    "COMPLETED": "DCFCE7", "REPORT_SHARED": "DCFCE7", "SCHEDULED": "DBEAFE",
    "RESCHEDULED": "DBEAFE", "QUOTED": "FEF3C7", "INFO_REQUESTED": "FEF3C7",
    "ACK_WORKING": "F3E8FF", "CANCELLED_OR_NEGATIVE": "FEE2E2", "OTHER": "F3F4F6",
}
if handling and len(handling) > 1:
    ws = wb.create_sheet("Handling")
    ws.freeze_panes = "A2"
    hh = handling[0]
    for i, h in enumerate(hh, 1):
        ws.cell(row=1, column=i, value=h.replace("_", " ").title())
    style_header(ws, 1, len(hh))
    ridx = {h: i for i, h in enumerate(hh)}
    for ri, row in enumerate(handling[1:], start=2):
        for i, v in enumerate(row, 1):
            ws.cell(row=ri, column=i, value=v)
        res = row[ridx.get("resolution", -1)] if "resolution" in ridx else ""
        if res in RES_FILL:
            ws.cell(row=ri, column=ridx["resolution"] + 1).fill = PatternFill("solid", fgColor=RES_FILL[res])
    hw = {"group": 24, "intent": 15, "ids": 12, "question": 52, "team_reply": 52,
          "replied_by": 20, "resolution": 20, "latency_min": 11, "matched_by": 12}
    for h, w in hw.items():
        if h in ridx:
            ws.column_dimensions[get_column_letter(ridx[h] + 1)].width = w
    last = get_column_letter(len(hh)) + str(len(handling))
    tbl = Table(displayName="Handling", ref=f"A1:{last}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight11", showRowStripes=True)
    ws.add_table(tbl)

# ══════════════════════════ ReplyLibrary (learned) ═══════════════════════
if reply_library:
    ws = wb.create_sheet("ReplyLibrary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "What the team actually replies (learned from real traffic)"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = "Per question intent: how it gets resolved, and the real phrasings the team uses (# = a number/id). These become the bot's grounded reply templates."
    ws["A2"].font = Font(size=10, color=MUTED)
    rl = ["Intent", "Pairs", "Resolution mix", "Top real reply (learned)", "Seen"]
    for i, h in enumerate(rl, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(rl))
    r = 5
    for intent, v in sorted(reply_library.items(), key=lambda kv: -kv[1]["pairs"]):
        mix = ", ".join(f"{k} {n}" for k, n in list(v["resolutions"].items())[:4])
        tops = v["topReplies"][:6] or [{"template": "—", "n": ""}]
        start = r
        for t in tops:
            ws.cell(row=r, column=4, value=t["template"])
            ws.cell(row=r, column=5, value=t["n"])
            for cc in range(1, 6):
                ws.cell(row=r, column=cc).border = BORDER
            r += 1
        ws.cell(row=start, column=1, value=intent).font = Font(bold=True)
        ws.cell(row=start, column=2, value=v["pairs"])
        ws.cell(row=start, column=3, value=mix)
        if r - start > 1:
            for col in (1, 2, 3):
                ws.merge_cells(start_row=start, start_column=col, end_row=r - 1, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in {"A": 16, "B": 7, "C": 40, "D": 60, "E": 7}.items():
        ws.column_dimensions[col].width = w

path = os.path.join(OUT, "LabStack-WA-Bot-Actions.xlsx")
wb.save(path)
print("wrote", path)
print("Messages rows:", len(ROWS))
